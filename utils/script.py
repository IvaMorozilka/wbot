from utils.excel_parser_utils import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from icecream import ic

ic.configureOutput(prefix="LOGS| ")
ic.disable()


def transform_pipeline(input, output):
    list_of_warnings = []
    # Проверка наличия активного листа

    workbook = load_workbook(input, data_only=True, read_only=False)
    sheet = workbook.active

    if sheet is None:
        return False, "Файл не содержит активного листа."

    if sheet.max_row == 0 or sheet.max_column == 0:
        return False, "Активный лист пустой."

    # Разъеденим все объедененные ячейки, они мешают.
    sheet = unmerge_cells(sheet)

    if sheet["A1"].value is None or sheet["A1"].value == "":
        list_of_warnings.append(
            "Данные не начинаются с ячейки A1. Было применено удаление строк. Рекомендуется проверить результат!"
        )

        first_non_empty = None
        for idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=1)):
            if row[0].value is not None:
                first_non_empty = idx + 1
                break
        if first_non_empty is not None and first_non_empty > 1:
            sheet.delete_rows(1, first_non_empty - 1)

    # Удаление столбца A
    process_and_delete_column(sheet, "A")

    # Удаление всех фильтров
    sheet.auto_filter.ref = None

    # Раскрываем скрытые строки и столбцы
    sheet = remove_hidden_cells(sheet)

    # Увеличим ширину столбцов, чтобы влезали цифры и было красиво
    set_column_width(sheet, ["B", "C", "D", "E", "F", "G", "H", "I"], 18)
    set_column_width(sheet, ["A"], 63)

    code, msg = process_header(sheet)
    if not code:
        return code, msg

    otvetst_col_letter = get_column_letter(sheet.max_column - 1)

    # Этап добавления рaсчетов
    # data[0] расчеты в руб, data[1] - в %
    code, msg, data, logs = calculate_additional_data(sheet)
    if not code:
        return False, msg

    for log in logs:
        if any(value == "" for value in log.values()):
            return (
                False,
                'При расчетах произошла ошибка. Не были найдены ячейки "Развитие" или "Сопровождение" в колонке "Наименование показателя эффективности и результативности деятельности учреждения". Возможно опечатки.',
            )

    if any(len(value.split("+")) < 6 for value in logs[0].values()):
        list_of_warnings.append(
            "Внимание. Настоятельно рекомендуется проверить расчеты с помощью логов. Не все суммы (руб) содержат 6 значений."
        )

    if any(len(value.split("/")[1].split("+")) < 6 for value in logs[1].values()):
        list_of_warnings.append(
            "Внимание. Настоятельно рекомендуется проверить расчеты с помощью логов. Не все выражения (проценты) содержат 6 значений в знаменателе."
        )

    last_row = find_last_row_with_word(sheet, otvetst_col_letter, "Бекетова") + 1
    col_rub = get_column_letter(find_column_index_by_header(sheet, ["Итого", "руб"]))
    col_perc = get_column_letter(find_column_index_by_header(sheet, ["Итого", "%"]))

    sheet.insert_rows(last_row, 6)
    for key, value_rub, value_perc in zip(
        data[0].keys(), data[0].values(), data[1].values()
    ):
        sheet[f"A{last_row}"].value = key
        sheet[f"{col_rub}{last_row}"] = value_rub
        sheet[f"{col_rub}{last_row}"].number_format = "0.00"
        sheet[f"{otvetst_col_letter}{last_row}"].value = "Бекетова"
        sheet[f"{col_perc}{last_row}"] = value_perc
        sheet[f"{col_perc}{last_row}"].number_format = "0.00%"

        last_row += 1

    # Переносим Гречушкина выше
    move_and_replace_rows(sheet, otvetst_col_letter, "Гречушкин", last_row)

    # Заполняем id
    # Пока не будем удалять строку, чтобы было легче анализировать логи по вычислениям
    # sheet.delete_rows(2)
    delete_empty_rows(sheet)
    fill_column_with_ids(sheet, 2, 2, get_column_letter(sheet.max_column))
    sheet = replace_bad_values(sheet)
    apply_borders_to_all_cells(sheet)
    apply_font_to_all_cells(sheet, "Times New Roman", 11)

    try:
        workbook.save(output)
        return True, "Файл успешно обработан!", list_of_warnings, logs

    except Exception as e:
        return True, f"Произошла ошибка при сохранении файла. {e}"


def process_document_by_option(input: str, output: str, category: str):
    if category == "ОЭП":
        return "ОБРАБОТАЛ ДОКУМЕНТ ОЭП"
    else:
        return "Обработал другие документы"
