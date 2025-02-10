from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, Side
from icecream import ic
import re

# Определяем список функций, которые будут доступны при импорте через `from utils import *`
__all__ = [
    "unmerge_cells",
    "remove_hidden_cells",
    "set_column_width",
    "replace_bad_values",
    "move_and_replace_rows",
    "fill_column_with_ids",
    "apply_font_to_all_cells",
    "apply_borders_to_all_cells",
    "delete_empty_rows",
    "is_number",
    "process_and_delete_column",
    "process_header",
    "find_last_row_with_word",
    "find_column_index_by_header",
    "calculate_additional_data",
]


def unmerge_cells(sheet):
    merged_cells = sheet.merged_cells.ranges
    merged_cells_ranges = (
        re.sub(r"(MergedCellRange)|([<>{}])", "", string=str(merged_cells))
        .replace(" ", "")
        .split(",")
    )
    for r in merged_cells_ranges:
        sheet.unmerge_cells(r)

    return sheet


def remove_hidden_cells(sheet):
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].hidden = False

    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].hidden = False

    return sheet


def set_column_width(sheet, columns, width):
    for col in columns:
        sheet.column_dimensions[col].width = width


def replace_bad_values(sheet, remove_control_char=False):
    # Замена значений в ячейках
    def remove_control_characters(text):
        if text is None:
            return text
        # Регулярное выражение для удаления управляющих символов и пробелов
        # Удаляем все управляющие символы
        text = re.sub(r"[\x00-\x1F\x7F]", "", str(text))
        # Удаляем пробелы в начале строки
        text = re.sub(r"^\s+", "", text)
        return text

    for row in sheet.iter_rows():
        for cell in row:
            # Заменяем "-" на пустое значение
            if cell.value == "-":
                cell.value = ""
            # Замена ошибки деления 0 на 0
            if cell.data_type == "e" and cell.value == "#DIV/0!":
                cell.value = "0,00%"
            # Замена 90-100 на 100 и тому подобных, на правое значение
            if cell.value != "" and cell.data_type == "s":
                if all(symb in cell.value for symb in ["-", "%"]):
                    cell.value = cell.value.split("-")[1]

    if remove_control_char:
        for row in sheet.iter_rows(min_col=1, max_col=1):  # Только первая колонка
            for cell in row:
                cell.value = remove_control_characters(cell.value)
    return sheet


def move_and_replace_rows(sheet, search_column, search_value, target_row_index):
    # Преобразуем букву столбца в числовой индекс
    column_index = column_index_from_string(search_column)

    # Список для хранения строк, которые нужно переместить
    rows_to_move = []

    # Сначала собираем все строки, которые нужно переместить
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if (
            row[column_index - 1] == search_value
        ):  # Проверяем значение в указанном столбце
            rows_to_move.append(row)

    # Если строк для перемещения нет, выходим из функции
    if not rows_to_move:
        print("Строки с указанным значением не найдены.")
        return

    # Удаляем найденные строки из их текущего места
    for row in reversed(range(1, sheet.max_row + 1)):
        if sheet.cell(row=row, column=column_index).value == search_value:
            sheet.delete_rows(row)

    # Вставляем строки на новое место
    sheet.insert_rows(target_row_index, amount=len(rows_to_move))

    # Записываем данные в новые строки
    for i, row in enumerate(rows_to_move, start=target_row_index):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)


def fill_column_with_ids(sheet, start_id, start_row, column):
    column_index = column_index_from_string(column)

    # Определяем последнюю строку с данными
    last_row = sheet.max_row
    for row in range(sheet.max_row, 0, -1):
        if any(
            sheet.cell(row=row, column=col).value
            for col in range(1, sheet.max_column + 1)
        ):
            last_row = row
            break

    # Заполняем столбец ID
    current_id = start_id
    for row in range(start_row, last_row + 1):
        sheet.cell(row=row, column=column_index, value=current_id)
        current_id += 1


def apply_font_to_all_cells(sheet, font_name="Calibri", font_size=11):
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)


def apply_borders_to_all_cells(sheet, border_style="thin"):
    side = Side(border_style=border_style)
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border


def delete_empty_rows(sheet):
    # Проходим по строкам с конца
    for row in range(sheet.max_row, 0, -1):
        if (
            sheet.cell(row=row, column=1).value is None
            or sheet.cell(row=row, column=1).value == ""
        ):
            sheet.delete_rows(row)
        # if sheet.cell(row=row, column=1).value is not None or sheet.cell(row=row, column=1).value != "":
        #     break


def is_number(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        # Удаляем точку и проверяем, состоит ли оставшаяся строка только из цифр
        cleaned_value = value.replace(".", "", 1)
        return cleaned_value.isdigit()
    return False


def process_and_delete_column(sheet, column_letter):
    # Преобразуем букву столбца в числовой индекс
    column_index = column_index_from_string(column_letter)

    # Проходим по всем строкам в столбце
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        right_cell = sheet.cell(row=row, column=column_index + 1)

        if cell.value == "" or cell.value is None:
            continue

        # Если в ячейке текст, переносим его в соседнюю правую ячейку
        if not is_number(cell.value):
            right_cell.value = cell.value
            cell.value = None  # Очищаем текущую ячейку

    # Удаляем столбец после обработки
    sheet.delete_cols(column_index)


def process_header(sheet):
    # Находим индексы ключевых ячеек
    responsible_col = None
    itogo_col = None
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if "ответ" in cell_value.lower():
            responsible_col = col
        elif "итого" in cell_value.lower():
            itogo_col = col

    # Если не найдены ключевые ячейки, выходим
    if responsible_col is None or itogo_col is None:
        return False, "Не найдены ячейки 'Ответственный' или 'Итого'."

    # Добавляем ячейку "ID" справа от "Ответственный", если справа пусто
    if (
        responsible_col + 1 > sheet.max_column
        or sheet.cell(row=1, column=responsible_col + 1).value is None
    ):
        sheet.insert_cols(responsible_col + 1)
        sheet.cell(row=1, column=responsible_col + 1, value="ID")

    # Оставляем только 4 столбца между 3-й ячейкой и "Итого"
    start_col = 2  # Начинаем с 3-й ячейки
    end_col = itogo_col - 1  # Заканчиваем перед "Итого"
    total_columns = end_col - start_col  # Общее количество столбцов

    if total_columns >= 4:
        columns_to_keep = 4
    else:
        columns_to_keep = 2

    # Вычисляем, сколько столбцов нужно удалить
    columns_to_delete = total_columns - columns_to_keep

    if columns_to_delete > 0:
        # Удаляем лишние столбцы слева от "Итого"
        sheet.delete_cols(start_col, columns_to_delete)

    return True, ""

    # if columns_to_keep == 2:
    #      for col, repl in zip(range(itogo_col - columns_to_keep, itogo_col), ['текущий', 'текущий']):
    #         cell_value = sheet.cell(row=1, column=col).value
    #         words = cell_value.split()
    #         words[1] = repl
    #         sheet.cell(row=1, column=col).value = " ".join(words)
    # else:
    #     for col, repl in zip(range(itogo_col - columns_to_keep, itogo_col), ['предыдущий', 'предыдущий', 'текущий', 'текущий']):
    #         cell_value = sheet.cell(row=1, column=col).value
    #         words = cell_value.split()
    #         words[1] = repl
    #         sheet.cell(row=1, column=col).value = " ".join(words)


def find_last_row_with_word(sheet, column_letter, search_word):
    # Преобразуем букву столбца в числовой индекс
    column_index = column_index_from_string(column_letter)

    # Проходим по строкам с конца
    for row in range(sheet.max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value == search_word:
            return row
    # Если слово не найдено
    return None


def find_column_index_by_header(sheet, header_symbols):
    # Проходим по всем столбцам в первой строке
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if all(symb in cell_value for symb in header_symbols):
            return col

    # Если заголовок не найден
    return None


def calculate_additional_data(sheet):
    col_rub = get_column_letter(find_column_index_by_header(sheet, ["Итого", "руб"]))
    col_perc = get_column_letter(find_column_index_by_header(sheet, ["Итого", "%"]))

    if col_rub is None or col_perc is None:
        return False, "Не найдены столбцы 'Итого руб' или 'Итого %'."

    prin_pos_start = find_last_row_with_word(
        sheet, "A", "Принято бюджетных обязательств"
    )
    isp_pos_start = find_last_row_with_word(
        sheet, "A", "Исполнено бюджетных обязательств"
    )
    vydel_pos_start = find_last_row_with_word(sheet, "A", "Выделенно бюджетных средств")
    prin_pos_end = find_last_row_with_word(
        sheet,
        "A",
        "Принято бюджетных обязательств (по месяцам нарастающим итогом) - ФАКТ",
    )
    isp_pos_end = find_last_row_with_word(
        sheet,
        "A",
        "Исполнено бюджетных обязательств (по месяцам нарастающим итогом) - ФАКТ",
    )
    vydel_pos_end = prin_pos_start

    if prin_pos_start is None or isp_pos_start is None or vydel_pos_start is None:
        return (
            False,
            "Не найдены ключевые строки ('Принято бюджетных обязательств' или 'Исполнено бюджетных обязательств') для расчетов.",
        )

    if prin_pos_end is None or isp_pos_end is None or vydel_pos_end is None:
        return (
            False,
            "Не найдены ключевые строки ('Принято бюджетных обязательств (по месяцам нарастающим итогом) - ФАКТ' или 'Исполнено бюджетных обязательств (по месяцам нарастающим итогом) - ФАКТ') для расчетов.",
        )

    # ic(prin_pos_start, prin_pos_end, isp_pos_start, isp_pos_end)

    content = [
        "Развитие ИС Принято",
        "Развитие ИС исполнено",
        "Сопровождение Принято",
        "Сопровождение Исполнено",
        "ИС Принято",
        "ИС Исполнено",
    ]
    logs_rub = dict.fromkeys(content, "")
    logs_perc = dict.fromkeys(content, "")
    result_rub = dict.fromkeys(content, 0)
    result_perc = dict.fromkeys(content, 0)
    result_temp = [0, 0, 0]

    # Шаблоны для поиска. Так как появились опечатки в док-ах 25 года. Неизвестно когда их не станет (:
    razvitie = ["развитие", "развтие"]
    soprovogdenie = ["сопровождение"]

    for i in range(prin_pos_start, prin_pos_end + 1):
        if str(sheet[f"A{i}"].value).strip().lower() in razvitie:
            logs_rub["Развитие ИС Принято"] += f"{col_rub}{i} "
            logs_rub["ИС Принято"] += f"{col_rub}{i - 1} "

            result_rub["Развитие ИС Принято"] += sheet[f"{col_rub}{i}"].value
            result_rub["ИС Принято"] += sheet[f"{col_rub}{i - 1}"].value

        elif str(sheet[f"A{i}"].value).strip().lower() in soprovogdenie:
            logs_rub["Сопровождение Принято"] += f"{col_rub}{i} "
            result_rub["Сопровождение Принято"] += sheet[f"{col_rub}{i}"].value

    for i in range(isp_pos_start, isp_pos_end + 1):
        if str(sheet[f"A{i}"].value).strip().lower() in razvitie:
            logs_rub["Развитие ИС исполнено"] += f"{col_rub}{i} "
            logs_rub["ИС Исполнено"] += f"{col_rub}{i - 1} "

            result_rub["Развитие ИС исполнено"] += sheet[f"{col_rub}{i}"].value
            result_rub["ИС Исполнено"] += sheet[f"{col_rub}{i - 1}"].value
        elif str(sheet[f"A{i}"].value).strip().lower() in soprovogdenie:
            logs_rub["Сопровождение Исполнено"] += f"{col_rub}{i} "
            result_rub["Сопровождение Исполнено"] += sheet[f"{col_rub}{i}"].value

    for i in range(vydel_pos_start, vydel_pos_end + 1):
        if str(sheet[f"A{i}"].value).strip().lower() in razvitie:
            logs_perc["Развитие ИС исполнено"] += f"{col_rub}{i} "
            logs_perc["Развитие ИС Принято"] += f"{col_rub}{i} "
            logs_perc["ИС Исполнено"] += f"{col_rub}{i - 1} "
            logs_perc["ИС Принято"] += f"{col_rub}{i - 1} "

            result_temp[0] += sheet[f"{col_rub}{i}"].value
            result_temp[2] += sheet[f"{col_rub}{i - 1}"].value
        elif str(sheet[f"A{i}"].value).strip().lower() in soprovogdenie:
            logs_perc["Сопровождение Исполнено"] += f"{col_rub}{i} "
            logs_perc["Сопровождение Принято"] += f"{col_rub}{i} "
            result_temp[1] += sheet[f"{col_rub}{i}"].value

    result_temp = [x for x in result_temp for _ in range(2)]

    for i, key in enumerate(result_perc.keys()):
        logs_perc[key] = f"ИТОГ{key.replace(' ', '')}/{logs_perc[key]}"
        try:
            result_perc[key] = result_rub[key] / result_temp[i]
        except ZeroDivisionError:
            result_perc[key] = 0

    logs = [logs_rub, logs_perc]

    logs[0] = {key: value.rstrip().replace(" ", "+") for key, value in logs[0].items()}
    logs[1] = {key: value.rstrip().replace(" ", "+") for key, value in logs[1].items()}

    ic(logs)
    result = [result_rub, result_perc]
    # ic(result)

    return True, "", result, logs
