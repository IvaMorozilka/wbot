import openpyxl
from openpyxl.utils import get_column_letter
from icecream import ic
import pandas as pd
import io
from create_bot import logger

from utils.excel_helpers.constants import (
    COLUMNS_DATA_TYPES,
    CORRECT_HEADER_DATA,
    COLUMNS_TO_CHECK_NULLS_IN,
)

ic.configureOutput(prefix="LOGS| ")
ic.disable()


class DocumentCheckError(Exception):
    pass


def check_correct_header(
    df: pd.DataFrame, category: str, correct_headers: dict = CORRECT_HEADER_DATA
):
    correct_headers_list = correct_headers.get(category, [])
    if not correct_headers_list:
        return False, "Остутсвуют правильные заголовки"
    xls_headers = [header.strip().lower() for header in df.columns]
    missing = {}

    for i, correct_header in enumerate(correct_headers_list):
        if correct_header.lower() not in xls_headers:
            missing[get_column_letter(i + 1)] = correct_header

    if missing:
        missing = [
            f"<b>{column}</b>: <code>{name}</code>\n"
            for column, name in missing.items()
        ]
        return (
            False,
            f"<b>Остуствуют столбцы</b> ⬇️\n{''.join(missing)}",
        )

    return True, ""


def check_merged_cells(file_bytes: io.BytesIO):
    wb = openpyxl.load_workbook(file_bytes)
    if wb.active.merged_cells:
        return False, "<b>❗️Найдены объединенные ячейки</b>"
    return True, ""


def check_missing_cells(
    df: pd.DataFrame, category: str, check_columns: dict = COLUMNS_TO_CHECK_NULLS_IN
) -> None:
    columns_to_check = check_columns.get(category, [])
    if not columns_to_check:
        columns_to_check = df.columns
    bad_columns = []

    for column in columns_to_check:
        if column in df.columns:
            series = df[column]
            if (
                series.isna().any() or (series == "-").any()
            ) and column in columns_to_check:
                bad_columns.append("<code>" + column + "</code>")

    if bad_columns:
        return (
            False,
            f"<b>❗️Наличие пропущенных значений</b> ⬇️\n{'\n'.join(bad_columns)}",
        )
    return True, ""


def check_data_types(
    df: pd.DataFrame, category: str, column_types: dict = COLUMNS_DATA_TYPES
) -> None:
    data_types = column_types.get(category, [])
    if not data_types:
        return False, "Не указаны типы данных"
    bad_columns = []

    for column, series in df.items():
        if series.dtype != data_types[column]:
            bad_columns.append("<code>" + column + "</code>")
    if bad_columns:
        return False, f"<b>❗️Неверный тип данных</b> ⬇️\n{'\n'.join(bad_columns)}"
    return True, ""


def check_document_by_category(
    file_bytes: io.BytesIO,
    category: str,
):
    df = pd.read_excel(file_bytes)
    checks = [
        (check_correct_header, [df, category]),
        (check_merged_cells, [file_bytes]),
        (check_missing_cells, [df, category]),
        (check_data_types, [df, category]),
    ]
    errors = []
    for idx, (check_fn, args) in enumerate(checks):
        is_valid, error_msg = check_fn(*args)
        if not is_valid:
            if idx == 0:
                errors.append(error_msg)
                break
            errors.append(error_msg)

    if errors:
        return False, f"<b>ℹ️Проверка документа не пройдена</b>\n{'\n'.join(errors)}"

    return True, ""
